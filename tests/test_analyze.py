import unittest
import openpyxl
from exceltp.analyzer import *
from exceltp.condition import *
from exceltp.config import *


def create_condition(config_string, cond_name, criteria_date):
    condition_conf = yaml.load(config_string)
    cond = DateCondition(cond_name, condition_conf[cond_name])
    cond.set_criteria_date(criteria_date)
    return cond


class AnalyzerTestCase(unittest.TestCase):

    def create_test_config(self):
        config_string = """
config_version: 0.1
analyzer:
    - group_a:
        - test:
            column_name: a
            column_type: date
            row_startline: 2
            condition: today_equal
            value: -5
    - group_b:
        - test:
            column_name: a
            column_type: date
            row_startline: 2
            condition: today_equal
            value: -3
"""
        config = Config(None)
        config.read_raw_data(config_string)
        return config

    def assign_group_cond(self, analyzer):
        group = ConditionGroup("group1", None)
        group2 = ConditionGroup("group2", None)

        config_string = """
        test:
            column_name: a
            column_type: date
            row_startline: 2
            condition: today_equal
            value: -3
"""
        date = datetime.datetime(2016, 5, 14)
        cond = create_condition(config_string, 'test', date)
        group.add_condition(cond)

        config_string2 = """
        test2:
            column_name: a
            column_type: date
            row_startline: 2
            condition: today_equal
            value: -5
"""
        date2 = datetime.datetime(2016, 5, 14)
        cond2 = create_condition(config_string2, 'test2', date2)
        group2.add_condition(cond2)

        analyzer.add_condition_group(group)
        analyzer.add_condition_group(group2)

    def test_config_group_count_from_text(self):
        config = self.create_test_config()
        analyzer = Analyzer(config.get_analyzer_conf())
        self.assertEqual(analyzer.count_condition_group(), 2)

    def test_analyze_fail_no_xlsdata(self):
        config = self.create_test_config()
        analyzer = Analyzer(config.get_analyzer_conf())
        verbose = False
        self.assertEqual(analyzer.analyze(verbose), False)

    def test_analyze_single_worksheet(self):
        analyzer = Analyzer(None)
        self.assign_group_cond(analyzer)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "worksheet1"
        ws1['A1'] = 'Due date'
        ws1['A2'] = datetime.datetime(2016, 5, 11)
        ws1['A3'] = datetime.datetime(2016, 5, 9)
        ws1['A4'] = datetime.datetime(2017, 6, 21)
        ws1['A5'] = datetime.datetime(2016, 5, 11)
        ws1['B1'] = 'Name'
        ws1['B2'] = 'Tony'
        ws1['B3'] = 'Mike'
        ws1['B4'] = 'Lee'
        ws1['B5'] = 'Kim'

        analyzer.set_excel_workbook(wb)

        verbose = False
        self.assertEqual(analyzer.analyze(verbose), True)
        self.assertEqual(analyzer.count_analyze_data(), 3)

    def test_analyze_multi_worksheet(self):
        analyzer = Analyzer(None)
        self.assign_group_cond(analyzer)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "worksheet1"
        ws1['A1'] = 'Due date'
        ws1['A2'] = datetime.datetime(2016, 5, 11)
        ws1['A3'] = datetime.datetime(2016, 5, 9)
        ws1['A4'] = datetime.datetime(2017, 6, 21)
        ws1['A5'] = datetime.datetime(2016, 5, 11)
        ws1['B1'] = 'Name'
        ws1['B2'] = 'Tony'
        ws1['B3'] = 'Mike'
        ws1['B4'] = 'Lee'
        ws1['B5'] = 'Kim'

        ws2 = wb.create_sheet()
        ws2.title = "worksheet2"
        ws2['A1'] = 'Due date'
        ws2['A2'] = datetime.datetime(2016, 5, 11)
        ws2['A3'] = datetime.datetime(2016, 5, 9)
        ws2['A4'] = datetime.datetime(2017, 6, 21)
        ws2['A5'] = datetime.datetime(2016, 5, 11)
        ws2['B1'] = 'Name'
        ws2['B2'] = 'Tony'
        ws2['B3'] = 'Mike'
        ws2['B4'] = 'Lee'
        ws2['B5'] = 'Kim'

        ws3 = wb.create_sheet()
        ws3.title = "worksheet3"

        analyzer.set_excel_workbook(wb)
        verbose = False
        self.assertEqual(analyzer.analyze(verbose), True)
        self.assertEqual(analyzer.count_analyze_data(), 6)
